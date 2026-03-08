from __future__ import annotations
import io
import csv
from fastapi import UploadFile, File

from openpyxl import load_workbook

from ..models import (
    ParseQuestionsRequest,
    ParseQuestionsResponse,
    ParsedQuestion,
    Highlight,
    ExtractedEntity,
)
from ..llm import call_llm
from ..core import create_router, Dependencies

router = create_router()

PARSE_SYSTEM_PROMPT = """You are a data analyst expert. Given a list of business questions, extract all measures, dimensions, and filters from each question.

Rules:
- A MEASURE is an aggregatable numeric value (e.g., "net sales", "units sold", "gross margin")
- A DIMENSION is a categorical attribute used for grouping (e.g., "region", "category", "store")
- A FILTER is a constraint or time range (e.g., "last month", "west region", "store 123")

For each question, identify every measure, dimension, and filter mention with its exact character positions (start, end) in the original text.

Also provide a deduplicated list of unique entities across all questions, with an inferred database column name for each.

Respond ONLY with JSON (no markdown fences) in this exact format:
{
  "parsed_questions": [
    {
      "original_text": "the question text",
      "highlights": [
        {"text": "net sales", "type": "measure", "start": 10, "end": 19}
      ]
    }
  ],
  "entities": [
    {"name": "Net Sales", "type": "measure", "inferred_column": "net_sales_amt", "source_questions": [0]}
  ]
}

Be precise with character positions. Zero-indexed. The "text" field must exactly match the substring at [start:end] in original_text."""


@router.post("/parse-questions", response_model=ParseQuestionsResponse, operation_id="parseQuestions")
async def parse_questions(request: ParseQuestionsRequest, ws: Dependencies.Client) -> ParseQuestionsResponse:
    """Parse business questions to extract measures, dimensions, and filters."""
    questions_text = "\n".join(f"{i}. {q}" for i, q in enumerate(request.questions))
    result = call_llm(PARSE_SYSTEM_PROMPT, f"Business questions:\n{questions_text}", ws)

    parsed_questions = [
        ParsedQuestion(
            original_text=pq["original_text"],
            highlights=[Highlight(**h) for h in pq["highlights"]],
        )
        for pq in result.get("parsed_questions", [])
    ]
    entities = [ExtractedEntity(**e) for e in result.get("entities", [])]
    return ParseQuestionsResponse(parsed_questions=parsed_questions, entities=entities)


@router.post("/upload-questions", response_model=ParseQuestionsResponse, operation_id="uploadQuestions")
async def upload_questions(file: UploadFile = File(...), ws: Dependencies.Client = None) -> ParseQuestionsResponse:  # type: ignore[assignment]
    """Upload xlsx/csv file of business questions and parse them."""
    content = await file.read()
    questions: list[str] = []

    if file.filename and file.filename.endswith(".xlsx"):
        wb = load_workbook(filename=io.BytesIO(content), read_only=True)
        sheet = wb.active
        if sheet:
            for row in sheet.iter_rows(values_only=True):
                for cell in row:
                    if cell and isinstance(cell, str) and cell.strip():
                        questions.append(cell.strip())
    else:
        text = content.decode("utf-8")
        reader = csv.reader(io.StringIO(text))
        for row in reader:
            for cell in row:
                if cell.strip():
                    questions.append(cell.strip())

    return await parse_questions(ParseQuestionsRequest(questions=questions), ws)
